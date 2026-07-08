"""Skill library: CRUD + batch import from markdown frontmatter or excel sheets.

Skills are pointers to externally-hosted markdown playbooks. We persist only
metadata (name/link/description/category/tags) — never the body. The frontend
fetches `link` via the proxy endpoint to preview the markdown live.
"""

from __future__ import annotations

import io
import re
from datetime import datetime
from typing import TypedDict
from urllib.parse import unquote, urlparse
from uuid import uuid4

from app.adapters.async_sqlite_store import AsyncSQLiteStore
from app.domain.models import Skill


class SkillError(RuntimeError):
    pass


class SkillImportSkipped(TypedDict, total=False):
    file: str
    row: str
    reason: str


class SkillImportResult(TypedDict):
    created: list[Skill]
    skipped: list[SkillImportSkipped]


SkillList = list[Skill]
StringList = list[str]
SkillImportFiles = list[tuple[str, bytes]]
SkippedSkillImports = list[SkillImportSkipped]
Frontmatter = dict[str, object]


_URL_EXT_RE = re.compile(r"\.(md|markdown|txt|html?|json|ya?ml)$", re.IGNORECASE)


def _derive_name_from_url(url: str) -> str:
    """Fall back name when caller only supplied a link. Mirrors the frontend
    paste-parser logic: last path segment, decoded, common doc extension stripped.
    """
    try:
        parsed = urlparse(url)
    except Exception:
        return url
    segs = [s for s in parsed.path.split("/") if s]
    if not segs:
        return parsed.netloc or url
    last = unquote(segs[-1])
    last = _URL_EXT_RE.sub("", last)
    return last or parsed.netloc or url


_FRONTMATTER_RE = re.compile(r"^---\s*\n(.*?)\n---\s*\n", re.DOTALL)


def _parse_frontmatter(text: str) -> Frontmatter:
    """Tolerant YAML-subset parser for skill frontmatter.

    Supports `key: value`, `key: [a, b]`, and block scalars on `key:` followed by
    `  - item` lines. Avoids the pyyaml dependency for this narrow use case.
    """
    m = _FRONTMATTER_RE.match(text)
    if not m:
        return {}
    block = m.group(1)
    data: Frontmatter = {}
    pending_list: StringList | None = None
    for raw in block.splitlines():
        line = raw.rstrip()
        if not line.strip():
            pending_list = None
            continue
        if pending_list is not None and line.startswith(("  -", "\t-", "- ")):
            item = line.lstrip(" \t-").strip().strip("'").strip('"')
            if item:
                pending_list.append(item)
            continue
        pending_list = None
        if ":" not in line:
            continue
        key, _, value = line.partition(":")
        key = key.strip()
        value = value.strip()
        if not value:
            pending_list = []
            data[key] = pending_list
            continue
        if value.startswith("[") and value.endswith("]"):
            inner = value[1:-1].strip()
            items = [p.strip().strip("'").strip('"') for p in inner.split(",") if p.strip()]
            data[key] = items
        else:
            data[key] = value.strip("'").strip('"')
    return data


def _to_str_list(v: object) -> StringList:
    if isinstance(v, list):
        return [str(x).strip() for x in v if str(x).strip()]
    if isinstance(v, str):
        return [p.strip() for p in v.split(",") if p.strip()]
    return []


class SkillService:
    def __init__(self, store: AsyncSQLiteStore):
        self.store = store

    async def list(self, *, search: str | None = None, category: str | None = None) -> SkillList:
        return await self.store.list_skills(search=search, category=category)

    async def list_categories(self) -> StringList:
        return await self.store.list_skill_categories()

    async def add_category(self, name: str) -> str:
        name = (name or "").strip()
        if not name:
            raise SkillError("category name is required")
        await self.store.add_skill_category(name)
        return name

    async def delete_category(self, name: str, *, force: bool = False) -> None:
        name = (name or "").strip()
        if not name:
            raise SkillError("category name is required")
        in_use = await self.store.delete_skill_category(name)
        if in_use > 0 and not force:
            # The user-defined row was removed, but skills still reference this
            # category — re-add the entry to avoid silently losing the group
            # while signalling the issue back to the caller.
            await self.store.add_skill_category(name)
            raise SkillError(
                f"category '{name}' is still used by {in_use} skill(s). "
                "Reassign or delete those skills first, or pass force=true."
            )

    async def get(self, skill_id: str) -> Skill | None:
        return await self.store.load_skill(skill_id)

    async def create(
        self,
        *,
        name: str,
        link: str,
        description: str | None = None,
        category: str | None = None,
        tags: StringList | None = None,
    ) -> Skill:
        name = (name or "").strip()
        link = (link or "").strip()
        if not name:
            raise SkillError("name is required")
        if not link:
            raise SkillError("link is required")
        now = datetime.now()
        skill = Skill(
            id=str(uuid4()),
            name=name,
            link=link,
            description=(description or "").strip() or None,
            category=(category or "").strip() or None,
            tags=tags or [],
            created_at=now,
            updated_at=now,
        )
        await self.store.save_skill(skill)
        return skill

    async def update(
        self,
        skill_id: str,
        *,
        name: str | None = None,
        link: str | None = None,
        description: str | None = None,
        category: str | None = None,
        tags: StringList | None = None,
    ) -> Skill:
        existing = await self.store.load_skill(skill_id)
        if existing is None:
            raise SkillError(f"skill not found: {skill_id}")
        updated = Skill(
            id=existing.id,
            name=(name.strip() if name is not None else existing.name),
            link=(link.strip() if link is not None else existing.link),
            description=(description.strip() if description else None)
            if description is not None
            else existing.description,
            category=(category.strip() if category else None)
            if category is not None
            else existing.category,
            tags=tags if tags is not None else existing.tags,
            created_at=existing.created_at,
            updated_at=datetime.now(),
        )
        await self.store.save_skill(updated)
        return updated

    async def delete(self, skill_id: str) -> None:
        await self.store.delete_skill(skill_id)

    async def import_markdown(self, files: SkillImportFiles) -> SkillImportResult:
        """Each file = one skill. Reads YAML frontmatter; expects `name` and
        `link` at minimum. Filename stem is used as a fallback name.
        """
        created: SkillList = []
        skipped: SkippedSkillImports = []
        for filename, raw in files:
            try:
                text = raw.decode("utf-8", errors="replace")
            except Exception as exc:
                skipped.append({"file": filename, "reason": f"decode error: {exc}"})
                continue
            meta = _parse_frontmatter(text)
            link = str(meta.get("link") or meta.get("url") or "").strip()
            if not link:
                skipped.append({"file": filename, "reason": "missing link/url in frontmatter"})
                continue
            name = str(meta.get("name") or "").strip()
            if not name:
                name = _derive_name_from_url(link) or filename.rsplit(".", 1)[0]
            description = str(meta.get("description") or "").strip() or None
            category = str(meta.get("category") or "").strip() or None
            tags = _to_str_list(meta.get("tags"))
            skill = await self.create(
                name=name,
                link=link,
                description=description,
                category=category,
                tags=tags,
            )
            created.append(skill)
        return {"created": created, "skipped": skipped}

    async def import_excel(self, content: bytes) -> SkillImportResult:
        """Reads .xlsx with header row mapping columns to fields. Recognised
        headers (case-insensitive): name, link/url, description, category, tags.
        Tags cell may be comma-separated.
        """
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise SkillError(
                "openpyxl is not installed — run `pip install openpyxl` in the backend venv"
            ) from exc
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return {"created": [], "skipped": []}
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return {"created": [], "skipped": []}
        col_map: dict[str, int] = {}
        for idx, cell in enumerate(header):
            if cell is None:
                continue
            key = str(cell).strip().lower()
            if key in ("url",):
                key = "link"
            col_map[key] = idx
        created: SkillList = []
        skipped: SkippedSkillImports = []
        for row_idx, row in enumerate(rows, start=2):

            def cell_value(field: str) -> str:
                idx = col_map.get(field)
                if idx is None or idx >= len(row):  # noqa: B023
                    return ""
                v = row[idx]  # noqa: B023
                return "" if v is None else str(v).strip()

            name = cell_value("name")
            link = cell_value("link")
            if not name and not link:
                continue
            if not link:
                skipped.append({"row": str(row_idx), "reason": "missing link"})
                continue
            if not name:
                name = _derive_name_from_url(link)
            if not name:
                skipped.append(
                    {"row": str(row_idx), "reason": "missing name and link could not be parsed"}
                )
                continue
            tags = _to_str_list(cell_value("tags"))
            skill = await self.create(
                name=name,
                link=link,
                description=cell_value("description") or None,
                category=cell_value("category") or None,
                tags=tags,
            )
            created.append(skill)
        return {"created": created, "skipped": skipped}
